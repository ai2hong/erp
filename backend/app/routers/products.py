"""
상품 API 라우터
GET  /products              — 전체 상품 목록 (매장별 재고 포함)
GET  /products/search       — 상품 검색
GET  /products/{id}         — 상품 단건
POST /products/upload-excel — 엑셀 일괄 등록/업데이트 [관리자+]
"""
import io
from typing import Annotated, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.core.deps import get_current_staff, RequireAdmin
from app.database import get_db
from app.models.product import Product, ProductCategory, SaleStatus
from app.models.staff import Staff

router = APIRouter()


@router.get("")
async def get_products(
    store_id: int = Query(..., description="매장 ID"),
    category: Optional[str] = Query(None),
    sale_status: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_staff=Depends(get_current_staff),
):
    q = select(Product).options(selectinload(Product.inventories))

    if category:
        q = q.where(Product.category == category)
    if sale_status:
        q = q.where(Product.sale_status == sale_status)
    else:
        q = q.where(Product.sale_status != SaleStatus.단종)

    q = q.order_by(Product.category, Product.name)
    result = await db.execute(q)
    products = result.scalars().all()

    return [
        {
            "id": p.id,
            "category": p.category,
            "name": p.name,
            "normal_price": p.normal_price,
            "device_discount_price": p.device_discount_price,
            "sale_status": p.sale_status,
            "stock": next(
                (
                    {
                        "qty_actual": i.qty_actual,
                        "qty_available": i.qty_available,
                        "is_shortage": i.is_shortage,
                        "is_out_of_stock": i.is_out_of_stock,
                    }
                    for i in p.inventories if i.store_id == store_id
                ),
                None,
            ),
        }
        for p in products
    ]


DEVICE_CATEGORIES = ["입호흡 기기", "입호흡 기기(단일가)", "폐호흡 기기", "폐호흡 기기(단일가)"]

@router.get("/search")
async def search_products(
    q: str = Query(..., min_length=1),
    device_only: bool = Query(False),
    db: AsyncSession = Depends(get_db),
    current_staff=Depends(get_current_staff),
):
    stmt = select(Product).where(
        Product.name.contains(q),
        Product.sale_status != SaleStatus.단종,
    )
    if device_only:
        stmt = stmt.where(Product.category.in_(DEVICE_CATEGORIES))
    result = await db.execute(stmt.order_by(Product.name).limit(20))
    products = result.scalars().all()
    return [{"id": p.id, "name": p.name, "category": p.category} for p in products]


VALID_CATEGORIES = {c.value for c in ProductCategory}
VALID_STATUSES   = {s.value for s in SaleStatus}
CATEGORY_LIST    = [c.value for c in ProductCategory]


@router.get("/excel-template")
async def download_excel_template(
    current_staff=Depends(get_current_staff),
):
    """품목 업로드용 엑셀 양식 다운로드"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "품목 업로드"

    # ── 헤더 스타일 ──
    hdr_fill   = PatternFill("solid", fgColor="1E3A5F")
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    center     = Alignment(horizontal="center", vertical="center")
    thin_side  = Side(style="thin", color="AAAAAA")
    thin_border = Border(left=thin_side, right=thin_side, bottom=thin_side)

    HEADERS = ["분류", "상품명", "정상가", "기기할인가", "판매상태", "메모"]
    REQ     = [True, True, True, False, False, False]
    WIDTHS  = [28, 36, 14, 14, 12, 30]

    for col, (h, req, w) in enumerate(zip(HEADERS, REQ, WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=f"{h}{'  *필수' if req else ''}")
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    # ── 분류 드롭다운 유효성 검사 ──
    cat_formula = '"' + ",".join(CATEGORY_LIST) + '"'
    dv_cat = DataValidation(type="list", formula1=cat_formula, allow_blank=False,
                            showErrorMessage=True,
                            errorTitle="분류 오류",
                            error="목록에서 선택해주세요")
    ws.add_data_validation(dv_cat)
    dv_cat.sqref = "A2:A1000"

    # ── 판매상태 드롭다운 ──
    dv_status = DataValidation(type="list", formula1='"판매중,품절,단종"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.sqref = "E2:E1000"

    # ── 예시 데이터 ──
    EXAMPLES = [
        ("입호흡 일반",        "망고파인애플 100ml",   12000, "",    "판매중", ""),
        ("폐호흡 일반",        "쿨민트 30ml",          9000,  "",    "판매중", ""),
        ("입호흡 기기",        "보이핏 스펙터 100W",   55000, 48000, "판매중", "기기 연동 할인가 있음"),
        ("입호흡 기기(단일가)","프리즘 미니",          38000, "",    "판매중", ""),
        ("입호흡 코일",        "메쉬 0.2옴",           8000,  "",    "판매중", ""),
        ("악세사리",           "실리콘 캡 세트",        3000,  "",    "판매중", ""),
    ]

    ex_fill   = PatternFill("solid", fgColor="EEF4FF")
    num_align = Alignment(horizontal="right")

    for row_idx, row in enumerate(EXAMPLES, start=2):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = ex_fill
            cell.border = Border(
                left=thin_side, right=thin_side,
                top=Side(style="thin", color="D0D8E8"),
                bottom=Side(style="thin", color="D0D8E8"),
            )
            if col_idx in (3, 4):
                cell.alignment = num_align

    # ── 안내 시트 ──
    ws2 = wb.create_sheet("분류 목록")
    ws2.column_dimensions["A"].width = 32
    ws2.cell(row=1, column=1, value="사용 가능한 분류명 (복사하여 붙여넣기)").font = Font(bold=True)
    for i, cat in enumerate(CATEGORY_LIST, start=2):
        ws2.cell(row=i, column=1, value=cat)

    # ── 반환 ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''%ED%92%88%EB%AA%A9_%EC%97%85%EB%A1%9C%EB%93%9C_%EC%96%91%EC%8B%9D.xlsx"},
    )


@router.post("/upload-excel")
async def upload_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current: Annotated[Staff, RequireAdmin] = None,
):
    """
    엑셀 파일로 상품 일괄 등록/업데이트.
    헤더 행(1행) 스킵 후 2행부터 읽음.
    컬럼 순서: 분류 | 상품명 | 정상가 | 기기할인가 | 판매상태 | 메모
    - 상품명으로 기존 상품 매칭 → 있으면 UPDATE, 없으면 INSERT
    - 분류, 상품명, 정상가 는 필수
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "xlsx 또는 xls 파일만 업로드 가능합니다")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl 패키지가 설치되지 않았습니다")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(400, "엑셀 파일을 읽을 수 없습니다")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    inserted = 0
    updated  = 0
    skipped  = 0
    errors   = []

    # 기존 상품 name → Product 매핑 (캐시)
    existing_map: dict[str, Product] = {}
    result = await db.execute(select(Product))
    for p in result.scalars().all():
        existing_map[p.name] = p

    for row_idx, row in enumerate(rows, start=2):
        if not any(row):        # 빈 행 스킵
            continue

        raw_cat, raw_name, raw_price, raw_disc, raw_status, raw_memo = (
            (row[i] if i < len(row) else None) for i in range(6)
        )

        # ── 필수값 검증 ──
        category_str = str(raw_cat).strip() if raw_cat else ""
        name         = str(raw_name).strip() if raw_name else ""
        price_str    = str(raw_price).strip() if raw_price is not None else ""

        if not name:
            errors.append(f"{row_idx}행: 상품명 누락")
            skipped += 1
            continue
        if category_str not in VALID_CATEGORIES:
            errors.append(f"{row_idx}행 [{name}]: 분류 '{category_str}' 유효하지 않음")
            skipped += 1
            continue
        try:
            normal_price = int(float(str(raw_price))) if raw_price not in (None, "", " ") else None
            if normal_price is None:
                raise ValueError
        except ValueError:
            errors.append(f"{row_idx}행 [{name}]: 정상가 오류")
            skipped += 1
            continue

        # ── 선택값 파싱 ──
        try:
            device_discount = int(float(str(raw_disc))) if raw_disc not in (None, "", " ") else None
        except ValueError:
            device_discount = None

        status_str = str(raw_status).strip() if raw_status else "판매중"
        if status_str not in VALID_STATUSES:
            status_str = "판매중"

        memo = str(raw_memo).strip() if raw_memo not in (None, "", " ") else None

        # ── Upsert ──
        if name in existing_map:
            p = existing_map[name]
            p.category             = category_str
            p.normal_price         = normal_price
            p.device_discount_price = device_discount
            p.sale_status          = status_str
            p.memo                 = memo
            updated += 1
        else:
            p = Product(
                category              = category_str,
                name                  = name,
                normal_price          = normal_price,
                device_discount_price = device_discount,
                sale_status           = status_str,
                memo                  = memo,
            )
            db.add(p)
            existing_map[name] = p
            inserted += 1

    await db.commit()

    return {
        "inserted": inserted,
        "updated":  updated,
        "skipped":  skipped,
        "errors":   errors[:50],   # 최대 50개만 반환
    }


@router.get("/{product_id}")
async def get_product(
    product_id: int,
    db: AsyncSession = Depends(get_db),
    current_staff=Depends(get_current_staff),
):
    result = await db.execute(select(Product).where(Product.id == product_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="상품을 찾을 수 없습니다")
    return p
